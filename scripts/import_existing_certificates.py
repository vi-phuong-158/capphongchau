"""Kiểm tra và nhập dữ liệu GCN đã được phường xác minh.

Mặc định là dry-run và chỉ sinh báo cáo không chứa CCCD/họ tên. Thêm ``--apply`` để ghi
Google Sheets. Yêu cầu Python 3.11+ và openpyxl; đọc cấu hình từ môi trường hoặc .env.local.
"""

from __future__ import annotations

import argparse
import hashlib
import hmac
import json
import os
import re
import sys
import unicodedata
import urllib.parse
import urllib.request
import uuid
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_PATTERN = "*11-11-2025.xlsx"
BACKFILL_VERSION = "CCCD_ONLY_V1"
NAMESPACE = uuid.UUID("867331b1-6210-5bfd-9f3c-2bcc2c03a345")


@dataclass(frozen=True)
class ValidRow:
    source_row: int
    issue_number: str
    issue_normalized: str
    issue_date: str
    registry_number: str
    full_name: str
    citizen_id: str
    role: str


def load_dotenv() -> None:
    path = ROOT / ".env.local"
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def normalized_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalized_issue(value: Any) -> str:
    return re.sub(r"[^0-9A-Z]", "", normalized_text(value).upper())


def iso_date(value: Any) -> str | None:
    parsed: date | None = None
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    else:
        text = normalized_text(value)
        for pattern in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(text, pattern).date()
                break
            except ValueError:
                pass
    if parsed is None:
        return None
    if parsed.year < 1950 or parsed > date.today():
        return None
    return parsed.isoformat()


def normalize_name(value: str) -> str:
    return re.sub(r"\s+", " ", unicodedata.normalize("NFC", value).strip()).upper()


def digest(pepper: str, value: str) -> str:
    return hmac.new(pepper.encode(), value.encode(), hashlib.sha256).hexdigest()


def identity_hashes(pepper: str, row: ValidRow) -> tuple[str, str]:
    # Khóa khớp là HMAC của riêng CCCD (citizen_hash) — phải trùng byte-for-byte với `identityHmac`
    # bên app. NGÀY SINH đã bị loại khỏi khóa: kho có 87% ngày sinh chỉ ghi mỗi năm nên đưa vào là
    # trượt phần lớn. match_hash (CCCD|họ tên) chỉ dùng để chống trùng chủ trong cùng một GCN.
    citizen_hash = digest(pepper, row.citizen_id)
    identity_match = digest(pepper, "|".join((row.citizen_id, normalize_name(row.full_name))))
    return citizen_hash, identity_match


def chunks(values: list[list[str]], size: int = 500) -> Iterable[list[list[str]]]:
    for start in range(0, len(values), size):
        yield values[start : start + size]


def a1_column(index: int) -> str:
    result = ""
    current = index
    while current:
        current, remainder = divmod(current - 1, 26)
        result = chr(65 + remainder) + result
    return result


def request_json(url: str, *, method: str = "GET", token: str = "", body: Any = None) -> Any:
    payload = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
    headers = {"accept": "application/json"}
    if payload is not None:
        headers["content-type"] = "application/json; charset=utf-8"
    if token:
        headers["authorization"] = f"Bearer {token}"
    request = urllib.request.Request(url, data=payload, headers=headers, method=method)
    with urllib.request.urlopen(request, timeout=45) as response:
        return json.loads(response.read().decode("utf-8"))


def access_token() -> str:
    required = [
        "GOOGLE_DRIVE_CLIENT_ID",
        "GOOGLE_DRIVE_CLIENT_SECRET",
        "GOOGLE_DRIVE_REFRESH_TOKEN",
    ]
    missing = [key for key in required if not os.environ.get(key)]
    if missing:
        raise RuntimeError("Thiếu biến môi trường: " + ", ".join(missing))
    body = urllib.parse.urlencode(
        {
            "client_id": os.environ[required[0]],
            "client_secret": os.environ[required[1]],
            "refresh_token": os.environ[required[2]],
            "grant_type": "refresh_token",
        }
    ).encode()
    request = urllib.request.Request(
        "https://oauth2.googleapis.com/token",
        data=body,
        headers={"content-type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        return json.loads(response.read().decode())["access_token"]


def append_values(token: str, spreadsheet_id: str, range_name: str, rows: list[list[str]]) -> None:
    for part in chunks(rows):
        encoded = urllib.parse.quote(range_name, safe="!:'")
        url = (
            f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/"
            f"{encoded}:append?valueInputOption=RAW&insertDataOption=INSERT_ROWS"
        )
        request_json(url, method="POST", token=token, body={"values": part})


def get_values(token: str, spreadsheet_id: str, range_name: str) -> list[list[str]]:
    encoded = urllib.parse.quote(range_name, safe="!:'")
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{encoded}"
    payload = request_json(url, token=token)
    return payload.get("values", [])


def append_bucket_values(
    token: str,
    spreadsheet_id: str,
    buckets: dict[int, list[list[str]]],
) -> None:
    existing = get_values(token, spreadsheet_id, "PUBLIC_LOOKUP_INDEX!A2:IV")
    column_lengths = [0] * 256
    for row_index, row in enumerate(existing, start=1):
        for column_index, value in enumerate(row[:256]):
            if value not in (None, ""):
                column_lengths[column_index] = row_index
    data: list[dict[str, Any]] = []
    for bucket, rows in sorted(buckets.items()):
        column = a1_column(bucket + 1)
        start_row = column_lengths[bucket] + 2
        data.append(
            {
                "range": f"PUBLIC_LOOKUP_INDEX!{column}{start_row}",
                "majorDimension": "ROWS",
                "values": rows,
            }
        )
    if not data:
        return
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values:batchUpdate"
    request_json(
        url,
        method="POST",
        token=token,
        body={"valueInputOption": "RAW", "data": data},
    )


def read_source(path: Path) -> tuple[list[ValidRow], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    valid: list[ValidRow] = []
    invalid: list[dict[str, Any]] = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
        issue_number = normalized_text(cells[2] if len(cells) > 2 else "")
        issue_normalized = normalized_issue(issue_number)
        issue_date = iso_date(cells[3] if len(cells) > 3 else None)
        registry_number = normalized_text(cells[4] if len(cells) > 4 else "")
        full_name = normalized_text(cells[7] if len(cells) > 7 else "")
        citizen_id = re.sub(r"\D", "", normalized_text(cells[10] if len(cells) > 10 else ""))
        role = normalized_text(cells[13] if len(cells) > 13 else "")
        reasons: list[str] = []
        if not issue_normalized:
            reasons.append("MISSING_ISSUE_NUMBER")
        if not issue_date:
            reasons.append("INVALID_ISSUE_DATE")
        if not full_name:
            reasons.append("MISSING_OWNER_NAME")
        if not re.fullmatch(r"\d{12}", citizen_id):
            reasons.append("INVALID_CITIZEN_ID")
        if reasons:
            invalid.append({"sourceRow": row_number, "reasons": reasons})
            continue
        valid.append(
            ValidRow(
                row_number,
                issue_number,
                issue_normalized,
                issue_date,
                registry_number,
                full_name,
                citizen_id,
                role,
            )
        )
    workbook.close()
    return valid, invalid


def build_rows(valid: list[ValidRow], pepper: str, import_id: str) -> tuple[
    list[list[str]], list[list[str]], dict[int, list[list[str]]], int, int
]:
    groups: dict[str, list[ValidRow]] = defaultdict(list)
    for item in valid:
        groups[item.issue_normalized].append(item)
    certificates: list[list[str]] = []
    owners: list[list[str]] = []
    buckets: dict[int, list[list[str]]] = defaultdict(list)
    duplicate_rows = 0
    conflict_rows = 0
    now = datetime.now().astimezone().isoformat()
    for issue_normalized, group in groups.items():
        variants = {(item.issue_date, item.registry_number) for item in group}
        status = "VERIFIED" if len(variants) == 1 else "CONFLICT"
        if status == "CONFLICT":
            conflict_rows += len(group)
        first = group[0]
        record_id = str(uuid.uuid5(NAMESPACE, issue_normalized))
        certificates.append(
            [
                record_id,
                first.issue_number,
                issue_normalized,
                first.issue_date,
                first.registry_number,
                status,
                import_id,
                ",".join(str(item.source_row) for item in group),
                now,
                now,
            ]
        )
        seen_owners: set[str] = set()
        for item in group:
            citizen_hash, match_hash = identity_hashes(pepper, item)
            owner_key = f"{record_id}:{match_hash}"
            if owner_key in seen_owners:
                duplicate_rows += 1
                continue
            seen_owners.add(owner_key)
            link_id = str(uuid.uuid5(NAMESPACE, owner_key))
            owners.append(
                [
                    link_id,
                    record_id,
                    citizen_hash,
                    match_hash,
                    item.role,
                    status,
                    import_id,
                    str(item.source_row),
                    now,
                ]
            )
            if status == "VERIFIED":
                buckets[int(citizen_hash[:2], 16)].append(
                    [
                        json.dumps(
                            {
                                "kind": "EXISTING",
                                "citizenIdHmac": citizen_hash,
                                "identityMatchHmac": match_hash,
                                "existingRecordId": record_id,
                            },
                            separators=(",", ":"),
                        )
                    ]
                )
    return certificates, owners, buckets, duplicate_rows, conflict_rows


def latest_rows_by_id(rows: list[list[str]]) -> dict[str, list[str]]:
    return {row[0]: row for row in rows if row and row[0]}


def existing_index_pairs(rows: list[list[str]]) -> set[tuple[str, str]]:
    pairs: set[tuple[str, str]] = set()
    for row in rows:
        for value in row:
            try:
                item = json.loads(value)
                if item.get("kind") == "EXISTING" and item.get("citizenIdHmac") and item.get("existingRecordId"):
                    pairs.add((item["citizenIdHmac"], item["existingRecordId"]))
            except (TypeError, ValueError, json.JSONDecodeError):
                continue
    return pairs


def backfill_rows(
    certificates: list[list[str]],
    owners: list[list[str]],
    existing_certificates: list[list[str]],
    existing_owners: list[list[str]],
    existing_pairs: set[tuple[str, str]],
) -> tuple[list[list[str]], list[list[str]], dict[int, list[list[str]]], int]:
    latest_certificates = latest_rows_by_id(existing_certificates)
    # Bản import trước từng dùng identity_match_hmac có ngày sinh; link_id vì vậy không còn
    # là khóa ổn định sau khi chuyển sang CCCD-đơn. Dedupe backfill theo GCN + HMAC CCCD.
    existing_owner_keys = {
        (row[1], row[2])
        for row in existing_owners
        if len(row) > 2 and row[1] and row[2]
    }
    changed_statuses = {
        record_id
        for record_id, certificate in ((row[0], row) for row in certificates)
        if record_id in latest_certificates and latest_certificates[record_id][5] != certificate[5]
    }
    certificate_appends = [
        certificate
        for certificate in certificates
        if certificate[0] not in latest_certificates
        or latest_certificates[certificate[0]][5] != certificate[5]
    ]
    owner_appends = [
        owner
        for owner in owners
        if (owner[1], owner[2]) not in existing_owner_keys or owner[1] in changed_statuses
    ]
    buckets: dict[int, list[list[str]]] = defaultdict(list)
    new_index_entries = 0
    for owner in owners:
        citizen_hash, record_id, status = owner[2], owner[1], owner[5]
        pair = (citizen_hash, record_id)
        if status != "VERIFIED" or pair in existing_pairs:
            continue
        existing_pairs.add(pair)
        new_index_entries += 1
        buckets[int(citizen_hash[:2], 16)].append(
            [
                json.dumps(
                    {
                        "kind": "EXISTING",
                        "citizenIdHmac": citizen_hash,
                        "identityMatchHmac": owner[3],
                        "existingRecordId": record_id,
                    },
                    separators=(",", ":"),
                )
            ]
        )
    return certificate_appends, owner_appends, buckets, new_index_entries


def run_backfill(
    source: Path,
    source_sha: str,
    valid: list[ValidRow],
    invalid: list[dict[str, Any]],
    pepper: str,
    apply: bool,
) -> int:
    spreadsheet_id = os.environ.get("GOOGLE_SHEETS_SPREADSHEET_ID", "")
    if not spreadsheet_id:
        raise RuntimeError("Thiếu GOOGLE_SHEETS_SPREADSHEET_ID.")
    token = access_token()
    source_import_id = str(uuid.uuid5(NAMESPACE, source_sha))
    backfill_id = str(uuid.uuid5(NAMESPACE, f"BACKFILL:{BACKFILL_VERSION}:{source_sha}"))
    desired_certificates, desired_owners, _, duplicates, conflicts = build_rows(valid, pepper, backfill_id)
    previous_runs = get_values(token, spreadsheet_id, "EXISTING_IMPORT_RUNS!A2:L")
    if not any(row and row[0] == source_import_id and len(row) > 3 and row[3] == "COMPLETED" for row in previous_runs):
        raise RuntimeError("Không tìm thấy lần import COMPLETED của đúng tệp nguồn; từ chối backfill.")
    completed = any(row and row[0] == backfill_id and len(row) > 3 and row[3] == "COMPLETED" for row in previous_runs)
    existing_certificates = get_values(token, spreadsheet_id, "EXISTING_CERTIFICATES!A2:J")
    existing_owners = get_values(token, spreadsheet_id, "EXISTING_CERTIFICATE_OWNERS!A2:I")
    existing_index = get_values(token, spreadsheet_id, "PUBLIC_LOOKUP_INDEX!A2:IV")
    certificate_appends, owner_appends, buckets, new_index_entries = backfill_rows(
        desired_certificates,
        desired_owners,
        existing_certificates,
        existing_owners,
        existing_index_pairs(existing_index),
    )
    report = {
        "operation": "BACKFILL",
        "backfillVersion": BACKFILL_VERSION,
        "importId": backfill_id,
        "sourceImportId": source_import_id,
        "sourceFileName": source.name,
        "sourceSha256": source_sha,
        "totalRows": len(valid) + len(invalid),
        "validRows": len(valid),
        "invalidRows": len(invalid),
        "conflictRows": conflicts,
        "duplicateRows": duplicates,
        "certificateAppends": len(certificate_appends),
        "ownerAppends": len(owner_appends),
        "indexAppends": new_index_entries,
        "alreadyCompleted": completed,
        "invalid": invalid,
    }
    report_dir = ROOT / "reports"
    report_dir.mkdir(exist_ok=True)
    report_path = report_dir / f"existing-backfill-{source_sha[:12]}.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({**report, "invalid": f"{len(invalid)} rows", "report": str(report_path)}, ensure_ascii=False, indent=2))
    if completed:
        print("BACKFILL_NOOP: lần backfill này đã COMPLETED.")
        return 0
    if not apply:
        print("DRY-RUN: chưa ghi Google Sheets. Dùng --backfill --apply sau khi duyệt báo cáo.")
        return 0
    started = [
        backfill_id,
        source.name,
        source_sha,
        "STARTED",
        str(report["totalRows"]),
        str(report["validRows"]),
        str(report["invalidRows"]),
        str(conflicts),
        str(duplicates),
        report_path.name,
        datetime.now().astimezone().isoformat(),
        "",
    ]
    if not any(row and row[0] == backfill_id for row in previous_runs):
        append_values(token, spreadsheet_id, "EXISTING_IMPORT_RUNS!A:L", [started])
    if certificate_appends:
        append_values(token, spreadsheet_id, "EXISTING_CERTIFICATES!A:J", certificate_appends)
    if owner_appends:
        append_values(token, spreadsheet_id, "EXISTING_CERTIFICATE_OWNERS!A:I", owner_appends)
    if buckets:
        append_bucket_values(token, spreadsheet_id, buckets)
    completed_row = started.copy()
    completed_row[3] = "COMPLETED"
    completed_row[11] = datetime.now().astimezone().isoformat()
    append_values(token, spreadsheet_id, "EXISTING_IMPORT_RUNS!A:L", [completed_row])
    print("Đã backfill dữ liệu thiếu theo khóa CCCD; có thể chạy lại an toàn để xác nhận no-op.")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", nargs="?", type=Path)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--backfill", action="store_true")
    args = parser.parse_args()
    load_dotenv()
    source = args.source
    if source is None:
        matches = list((ROOT / "Tai lieu").glob(DEFAULT_PATTERN))
        if len(matches) != 1:
            raise RuntimeError("Không xác định duy nhất tệp Excel nguồn.")
        source = matches[0]
    source = source.resolve()
    source_sha = hashlib.sha256(source.read_bytes()).hexdigest()
    import_id = str(uuid.uuid5(NAMESPACE, source_sha))
    pepper = os.environ.get("DATA_HASH_PEPPER", "")
    if len(pepper) < 32:
        raise RuntimeError("DATA_HASH_PEPPER phải có ít nhất 32 ký tự.")
    valid, invalid = read_source(source)
    if args.backfill:
        return run_backfill(source, source_sha, valid, invalid, pepper, args.apply)
    certificates, owners, buckets, duplicates, conflicts = build_rows(valid, pepper, import_id)
    report = {
        "importId": import_id,
        "sourceFileName": source.name,
        "sourceSha256": source_sha,
        "totalRows": len(valid) + len(invalid),
        "validRows": len(valid),
        "invalidRows": len(invalid),
        "certificateRecords": len(certificates),
        "ownerLinks": len(owners),
        "duplicateRows": duplicates,
        "conflictRows": conflicts,
        "invalid": invalid,
    }
    report_dir = ROOT / "reports"
    report_dir.mkdir(exist_ok=True)
    report_path = report_dir / f"existing-import-{source_sha[:12]}.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({**report, "invalid": f"{len(invalid)} rows", "report": str(report_path)}, ensure_ascii=False, indent=2))
    if not args.apply:
        print("DRY-RUN: chưa ghi Google Sheets. Dùng --apply sau khi duyệt báo cáo.")
        return 0
    spreadsheet_id = os.environ.get("GOOGLE_SHEETS_SPREADSHEET_ID", "")
    if not spreadsheet_id:
        raise RuntimeError("Thiếu GOOGLE_SHEETS_SPREADSHEET_ID.")
    token = access_token()
    previous_runs = get_values(token, spreadsheet_id, "EXISTING_IMPORT_RUNS!A2:L")
    matching_runs = [row for row in previous_runs if row and row[0] == import_id]
    if any(len(row) > 3 and row[3] == "COMPLETED" for row in matching_runs):
        print("Nguồn này đã được nhập hoàn tất; không ghi trùng.")
        return 0
    if matching_runs:
        raise RuntimeError("Lần nhập cùng nguồn đang ở trạng thái dở dang. Kiểm tra các tab trước khi chạy lại.")
    started_at = datetime.now().astimezone().isoformat()
    run_base = [
        import_id,
        source.name,
        source_sha,
        "STARTED",
        str(report["totalRows"]),
        str(report["validRows"]),
        str(report["invalidRows"]),
        str(conflicts),
        str(duplicates),
        report_path.name,
        started_at,
        "",
    ]
    append_values(token, spreadsheet_id, "EXISTING_IMPORT_RUNS!A:L", [run_base])
    append_values(token, spreadsheet_id, "EXISTING_CERTIFICATES!A:J", certificates)
    append_values(token, spreadsheet_id, "EXISTING_CERTIFICATE_OWNERS!A:I", owners)
    append_bucket_values(token, spreadsheet_id, buckets)
    completed = run_base.copy()
    completed[3] = "COMPLETED"
    completed[11] = datetime.now().astimezone().isoformat()
    append_values(token, spreadsheet_id, "EXISTING_IMPORT_RUNS!A:L", [completed])
    print("Đã nhập dữ liệu và chỉ mục HMAC vào Google Sheets.")
    return 0

if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(f"IMPORT_FAILED: {error}", file=sys.stderr)
        raise SystemExit(1)
